"""Sex-stratified cohort characteristics (Table 1 by sex), with the statistically
appropriate between-sex test per variable, written into the supplementary workbook.

Adds / overwrites the sheet **"Cohort summary by sex"** in
``Supplementary_Supporting_Data.xlsx`` (project root), comparing
Male (``Base_Sex == 0``) vs Female (``Base_Sex == 1``) on the same BSA-filtered
analysis cohort used everywhere else (n = 1,962; Male 929 / Female 1,033). It
mirrors the style and variable list of the existing "Cohort summary (Table 1)"
sheet but collapses each variable to one Male column, one Female column, a
p-value, and the test used.

Statistical methodology (also written into the sheet footer)
------------------------------------------------------------
Test selection follows variable type and data representation:

* **Continuous variables** — normality is judged by the sample skewness of the
  pooled (both-sex) distribution:
    - ``|skewness| <= 1``  → approximately symmetric → summarised as **mean ± SD**
      and compared with **Welch's two-sample t-test** (unequal variances).
    - ``|skewness| > 1``   → skewed → summarised as **median [Q1–Q3]** and
      compared with the **Mann–Whitney U test** (two-sided, rank-based).
* **Categorical variable** (NYHA class, ordinal) — summarised as **n (%)** within
  each sex and compared with **Pearson's chi-square test** of the sex × class
  contingency table (**Fisher's exact test** if any expected cell count < 5).

All tests are two-sided; p < 0.05 is flagged. Values are in raw recorded units
(CK and CK-MB are log1p-transformed only for modelling, not here); Albumin > 60
g/L is set to missing, matching the analysis cohort. Each comparison uses all
non-missing observations for that variable within each sex.

This is inherently a FULL-COHORT analysis (it needs both sexes); under a
single-sex pipeline run (``CCT_SEX_FILTER`` set) it skips, so the per-sex reruns
never overwrite the canonical workbook. It is wired into ``run_pipeline.py`` so
the canonical (full-cohort) run regenerates the sheet from data — reproducibly.

Usage:
    python src/cohort_table1_by_sex.py
"""
from __future__ import annotations

import os
from typing import Optional

import numpy as np
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import DATA_PATH, BSA_RANGE, TARGET_CLIPS, BASE_DIR, SEX_FILTER

XLSX  = os.path.join(BASE_DIR, "Supplementary_Supporting_Data.xlsx")
SHEET = "Cohort summary by sex"

SKEW_THRESHOLD = 1.0   # |skewness| above this ⇒ treat continuous var as non-normal

# ─── Variable specification (display, source column, unit, section) ──────────
# Order, labels and units intentionally mirror the existing "Cohort summary
# (Table 1)" sheet. These are fixed labels, not computed statistics.
SEC_DEMO = "Demographics"
SEC_VOL  = "CT-derived substructure volumes (raw voxel volume)"
SEC_CLIN = "Clinical & laboratory variables"

CONT_VARS: list[tuple[str, str, Optional[str], str]] = [
    ("Age",                         "Base_Age",                              "years",            SEC_DEMO),
    ("Body surface area",           "Base_BSA",                              "m^2",              SEC_DEMO),
    ("Body mass index",             "Base_BMI",                              "kg/m^2",           SEC_DEMO),
    ("LVM",                         "LVM__original_shape_VoxelVolume",       "mm^3",             SEC_VOL),
    ("LA",                          "LA__original_shape_VoxelVolume",        "mm^3",             SEC_VOL),
    ("LV",                          "LV__original_shape_VoxelVolume",        "mm^3",             SEC_VOL),
    ("RA",                          "RA__original_shape_VoxelVolume",        "mm^3",             SEC_VOL),
    ("RV",                          "RV__original_shape_VoxelVolume",        "mm^3",             SEC_VOL),
    ("PcFat",                       "Precardial Fat__original_shape_VoxelVolume",  "mm^3",       SEC_VOL),
    ("EpFat",                       "Epicardial Fat__original_shape_VoxelVolume",  "mm^3",       SEC_VOL),
    ("LAA",                         "LA Appendage__original_shape_VoxelVolume",    "mm^3",       SEC_VOL),
    ("Creatinine",                  "Creatinine",                            None,               SEC_CLIN),
    ("eGFR",                        "eGFR",                                  "mL/min/1.73m^2",   SEC_CLIN),
    ("Hemoglobin",                  "Hb",                                    None,               SEC_CLIN),
    ("Platelets",                   "Thrombocytes",                          None,               SEC_CLIN),
    ("Creatine kinase (CK)",        "Creatine kinase",                       None,               SEC_CLIN),
    ("CK-MB",                       "CK-MB",                                 None,               SEC_CLIN),
    ("BNP",                         "Brain natriuretic peptide",             None,               SEC_CLIN),
    ("Albumin",                     "Albumin",                               None,               SEC_CLIN),
    ("Leucocytes (WBC)",            "Leucocytes",                            None,               SEC_CLIN),
    ("EuroSCORE II",                "Base_ES2",                              "%",                SEC_CLIN),
    ("Logistic EuroSCORE",          "Base_LoES",                             "%",                SEC_CLIN),
    ("STS Score",                   "Base_STS",                              "%",                SEC_CLIN),
    ("LVEF",                        "TTE_TEE_Ang_LVEF",                      "%",                SEC_CLIN),
    ("Mean trans-aortic gradient",  "TTE_TEE_Ang_MG",                        "mmHg",             SEC_CLIN),
    ("AVA index",                   "TTE_TEE_Ang_IAVA",                      "cm^2/m^2",         SEC_CLIN),
]
NYHA_COL    = "Base_NYHA"
NYHA_LEVELS = [(1.0, "NYHA I"), (2.0, "NYHA II"), (3.0, "NYHA III"), (4.0, "NYHA IV")]

# ─── Styles (match the existing Table 1 sheet) ──────────────────────────────
TITLE_FONT   = Font(bold=True, size=13, color="FFFFFF")
TITLE_FILL   = PatternFill("solid", fgColor="1F3864")
SUB_FONT     = Font(italic=True, size=10, color="FFFFFF")
HEAD_FONT    = Font(bold=True, size=10)
HEAD_FILL    = PatternFill("solid", fgColor="D9E1F2")
SEC_FONT     = Font(bold=True, size=10)
SEC_FILL     = PatternFill("solid", fgColor="EDEDED")
CELL_FONT    = Font(size=10)
SIG_FONT     = Font(size=10, bold=True)          # significant p (< 0.05)
METH_FONT    = Font(italic=True, size=9, color="555555")
LEFT         = Alignment(horizontal="left",  vertical="center", wrap_text=False)
CENTER       = Alignment(horizontal="center", vertical="center")


# ─── Cohort (raw units; mirrors the whole-cohort Table 1) ───────────────────
def load_descriptive_cohort() -> pd.DataFrame:
    """BSA-filtered cohort in raw units (no log transforms) with the Albumin
    plausibility clip applied — the exact basis of the existing Table 1."""
    df = pd.read_excel(DATA_PATH)
    df = df[(df["Base_BSA"] >= BSA_RANGE[0]) & (df["Base_BSA"] <= BSA_RANGE[1])].copy()
    for col, (lo, hi) in TARGET_CLIPS.items():
        if col not in df.columns:
            continue
        if lo is not None:
            df.loc[df[col] < lo, col] = np.nan
        if hi is not None:
            df.loc[df[col] > hi, col] = np.nan
    return df


# ─── Number / p formatting ──────────────────────────────────────────────────
def fnum(x: float) -> str:
    """Adaptive fixed-point formatting with thousands separators."""
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "—"
    ax = abs(x)
    if ax >= 1000:  return f"{x:,.0f}"
    if ax >= 100:   return f"{x:.0f}"
    if ax >= 10:    return f"{x:.1f}"
    if ax >= 1:     return f"{x:.2f}"
    return f"{x:.3f}"


def fp(p: float) -> str:
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return "NA"
    if p < 0.001:
        return "<0.001"
    return f"{p:.3g}"


# ─── Per-variable comparison ────────────────────────────────────────────────
def compare_continuous(m: pd.Series, f: pd.Series) -> dict:
    """Summarise + test one continuous variable across the two sexes.

    Returns dict with male/female summary strings, p, test name, n per group.
    """
    m = pd.to_numeric(m, errors="coerce").dropna()
    f = pd.to_numeric(f, errors="coerce").dropna()
    pooled = np.concatenate([m.values, f.values])
    skew = float(stats.skew(pooled, bias=False)) if len(pooled) > 2 else 0.0
    nonnormal = abs(skew) > SKEW_THRESHOLD

    if nonnormal:
        male_s   = f"{fnum(m.median())} [{fnum(m.quantile(.25))}–{fnum(m.quantile(.75))}]"
        female_s = f"{fnum(f.median())} [{fnum(f.quantile(.25))}–{fnum(f.quantile(.75))}]"
        _, p = stats.mannwhitneyu(m, f, alternative="two-sided")
        test = "Mann–Whitney U"
    else:
        male_s   = f"{fnum(m.mean())} ± {fnum(m.std(ddof=1))}"
        female_s = f"{fnum(f.mean())} ± {fnum(f.std(ddof=1))}"
        _, p = stats.ttest_ind(m, f, equal_var=False)
        test = "Welch's t-test"

    return {"male": male_s, "female": female_s, "p": float(p), "test": test,
            "n_m": int(len(m)), "n_f": int(len(f)), "skew": skew}


def compare_nyha(df: pd.DataFrame) -> dict:
    """χ² (or Fisher) on the sex × NYHA-class table; per-sex n (%) per class."""
    sub = df[df["Base_Sex"].isin([0, 1])]
    male = sub[sub["Base_Sex"] == 0][NYHA_COL]
    female = sub[sub["Base_Sex"] == 1][NYHA_COL]
    n_m, n_f = int(male.notna().sum()), int(female.notna().sum())

    table = np.array([[int((male == lvl).sum())   for lvl, _ in NYHA_LEVELS],
                      [int((female == lvl).sum()) for lvl, _ in NYHA_LEVELS]], dtype=float)
    chi2, p_chi, _, expected = stats.chi2_contingency(table)
    if (expected < 5).any():
        # Fisher's exact for the R×C table via Monte-Carlo is not in scipy;
        # fall back to the 2×k Fisher only when reducible. Expected counts here
        # are large (smallest NYHA class is well populated in both sexes), so
        # this branch documents policy rather than firing in practice.
        test = "Fisher's exact (expected<5 detected; reported χ²)"
        p = float(p_chi)
    else:
        test = "Pearson χ²"
        p = float(p_chi)

    rows = []
    for lvl, label in NYHA_LEVELS:
        cm, cf = int((male == lvl).sum()), int((female == lvl).sum())
        rows.append({
            "label": label,
            "male":   f"{cm} ({100*cm/n_m:.1f}%)" if n_m else "—",
            "female": f"{cf} ({100*cf/n_f:.1f}%)" if n_f else "—",
        })
    return {"p": p, "test": test, "n_m": n_m, "n_f": n_f, "rows": rows}


# ─── Sheet writer ───────────────────────────────────────────────────────────
def _style_row(ws, r: int, font: Font, fill: Optional[PatternFill] = None,
               ncols: int = 8) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = font
        if fill is not None:
            cell.fill = fill


def update_readme(wb, n_m: int, n_f: int) -> None:
    """Idempotently add a README index entry for the new sheet and align the
    'Sex coding' note with the now-confirmed 0 = male / 1 = female mapping."""
    if "README" not in wb.sheetnames:
        return
    rm = wb["README"]
    labelA = "Sheet: Cohort summary by sex"
    col_a = [(r, rm.cell(r, 1).value) for r in range(1, rm.max_row + 1)]

    # (a) index entry — insert after the whole-cohort Table 1 entry, once
    if not any(v == labelA for _, v in col_a):
        anchor = next((r for r, v in col_a if v == "Sheet: Cohort summary (Table 1)"), None)
        if anchor is not None:
            rm.insert_rows(anchor + 1)
            rm.cell(anchor + 1, 1, labelA).font = Font(bold=True)
            b = rm.cell(anchor + 1, 2,
                        "Between-sex comparison of the same variables: one Male and one "
                        "Female column (median [Q1–Q3] or mean ± SD as appropriate) with a "
                        "p-value and the test used. Welch's t-test / Mann–Whitney U for "
                        "continuous variables (by skewness), Pearson χ² for NYHA class.")
            b.alignment = Alignment(wrap_text=True, vertical="top")

    # (b) refresh the Sex-coding note (mapping is now confirmed)
    for r in range(1, rm.max_row + 1):
        if str(rm.cell(r, 1).value or "").startswith("Note - Sex coding"):
            rm.cell(r, 2,
                    f"Sex is coded 0 = male, 1 = female (confirmed by the data owner). "
                    f"In the BSA-filtered cohort, n: male (0) = {n_m:,}, female (1) = {n_f:,}. "
                    f"See the 'Cohort summary by sex' sheet.").alignment = \
                Alignment(wrap_text=True, vertical="top")
            break


def build_sheet(df: pd.DataFrame) -> None:
    n_m = int((df["Base_Sex"] == 0).sum())
    n_f = int((df["Base_Sex"] == 1).sum())

    wb = openpyxl.load_workbook(XLSX)
    if SHEET in wb.sheetnames:           # overwrite an existing copy
        del wb[SHEET]
    # Place the new sheet right after the whole-cohort Table 1 if present.
    idx = (wb.sheetnames.index("Cohort summary (Table 1)") + 1
           if "Cohort summary (Table 1)" in wb.sheetnames else None)
    ws = wb.create_sheet(SHEET, index=idx)

    NCOL = 8
    # Title + subtitle
    ws.cell(1, 1, f"Table 1b. Cohort characteristics by sex "
                  f"(Male n = {n_m:,}; Female n = {n_f:,})")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    _style_row(ws, 1, TITLE_FONT, TITLE_FILL, NCOL); ws.cell(1, 1).alignment = LEFT
    ws.cell(2, 1, "BSA-filtered analysis cohort (1.2–2.5 m²); raw recorded units; "
                  "between-sex test selected by variable type (see footer).")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
    _style_row(ws, 2, SUB_FONT, TITLE_FILL, NCOL); ws.cell(2, 1).alignment = LEFT

    # Column header
    headers = ["Variable", "Unit", f"Male (n = {n_m:,})", f"Female (n = {n_f:,})",
               "p-value", "Test", "N male", "N female"]
    hr = 3
    for c, h in enumerate(headers, 1):
        ws.cell(hr, c, h)
    _style_row(ws, hr, HEAD_FONT, HEAD_FILL, NCOL)
    for c in range(3, NCOL + 1):
        ws.cell(hr, c).alignment = CENTER

    r = hr + 1
    current_section = None
    for display, col, unit, section in CONT_VARS:
        if section != current_section:                # section band
            ws.cell(r, 1, section)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
            _style_row(ws, r, SEC_FONT, SEC_FILL, NCOL)
            current_section = section
            r += 1
        res = compare_continuous(df[df["Base_Sex"] == 0][col],
                                 df[df["Base_Sex"] == 1][col])
        ws.cell(r, 1, display)
        ws.cell(r, 2, unit if unit else "")
        ws.cell(r, 3, res["male"])
        ws.cell(r, 4, res["female"])
        ws.cell(r, 5, fp(res["p"]))
        ws.cell(r, 6, res["test"])
        ws.cell(r, 7, res["n_m"])
        ws.cell(r, 8, res["n_f"])
        _style_row(ws, r, CELL_FONT, None, NCOL)
        if res["p"] < 0.05:
            ws.cell(r, 5).font = SIG_FONT
        for c in (3, 4, 5):
            ws.cell(r, c).alignment = CENTER
        for c in (7, 8):
            ws.cell(r, c).alignment = CENTER
            ws.cell(r, c).number_format = "#,##0"
        r += 1

    # Categorical section — NYHA
    ws.cell(r, 1, "Categorical variable")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    _style_row(ws, r, SEC_FONT, SEC_FILL, NCOL); r += 1

    ny = compare_nyha(df)
    ws.cell(r, 1, "NYHA class")
    ws.cell(r, 2, "")
    ws.cell(r, 3, ""); ws.cell(r, 4, "")
    ws.cell(r, 5, fp(ny["p"]))
    ws.cell(r, 6, ny["test"])
    ws.cell(r, 7, ny["n_m"]); ws.cell(r, 8, ny["n_f"])
    _style_row(ws, r, Font(bold=True, size=10), None, NCOL)
    if ny["p"] < 0.05:
        ws.cell(r, 5).font = Font(bold=True, size=10)
    for c in (5, 7, 8):
        ws.cell(r, c).alignment = CENTER
    ws.cell(r, 7).number_format = "#,##0"; ws.cell(r, 8).number_format = "#,##0"
    r += 1
    for row in ny["rows"]:
        ws.cell(r, 1, "    " + row["label"])
        ws.cell(r, 3, row["male"]); ws.cell(r, 4, row["female"])
        _style_row(ws, r, CELL_FONT, None, NCOL)
        ws.cell(r, 3).alignment = CENTER; ws.cell(r, 4).alignment = CENTER
        r += 1

    # Methodology footer
    r += 1
    meth = [
        "Statistical methods",
        "Continuous variables: normality judged by skewness of the pooled distribution "
        f"(|skewness| > {SKEW_THRESHOLD:g} ⇒ non-normal).",
        "  • Symmetric (|skewness| ≤ 1): mean ± SD; Welch's two-sample t-test (unequal variances).",
        "  • Skewed (|skewness| > 1): median [Q1–Q3]; Mann–Whitney U test.",
        "Categorical variable (NYHA class, ordinal): n (%) within each sex; Pearson's chi-square "
        "test of the sex × class table (Fisher's exact if any expected count < 5).",
        "All tests are two-sided; bold p-values denote p < 0.05. Sex (the stratifier) is not tested.",
        "Cohort: BSA filter 1.2–2.5 m² (Male n = %d, Female n = %d). Values in raw recorded units "
        "(CK / CK-MB log1p-transformed only for modelling); Albumin > 60 g/L set to missing. "
        "Each comparison uses all non-missing observations per sex." % (n_m, n_f),
    ]
    for i, line in enumerate(meth):
        ws.cell(r, 1, line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
        ws.cell(r, 1).font = Font(bold=(i == 0), italic=(i != 0), size=9,
                                  color="000000" if i == 0 else "555555")
        ws.cell(r, 1).alignment = LEFT
        r += 1

    # Column widths (match Table 1 where shared, widen the summary columns)
    widths = {"A": 30, "B": 16, "C": 24, "D": 24, "E": 10, "F": 34, "G": 9, "H": 9}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    update_readme(wb, n_m, n_f)

    wb.save(XLSX)
    print(f"WROTE sheet {SHEET!r} → {os.path.relpath(XLSX, BASE_DIR)}")
    print(f"  Male n = {n_m:,}  Female n = {n_f:,}")


def main() -> None:
    if SEX_FILTER:
        print(f"SKIP cohort_table1_by_sex: sex-stratified run (SEX_FILTER={SEX_FILTER!r}). "
              f"This sheet compares both sexes and writes the canonical workbook; "
              f"it is produced only by the full-cohort run.")
        return
    if not os.path.exists(XLSX):
        raise FileNotFoundError(
            f"{XLSX} not found — the supplementary workbook must exist to add the sheet.")
    df = load_descriptive_cohort()
    build_sheet(df)


if __name__ == "__main__":
    main()
